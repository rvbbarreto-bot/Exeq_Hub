"""Importação versionada do Anexo B — Lista de Serviço Nacional (XLSX)."""

from __future__ import annotations

from pathlib import Path

from django.db import transaction
from django.utils import timezone

from apps.master_data.models import NationalServiceCatalogVersion, NationalServiceItem


class NationalServiceImportError(ValueError):
    pass


def _norm_code(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def _int_field(value, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _lc116_hint(item: int, subitem: int) -> str:
    if item <= 0:
        return ""
    if subitem <= 0:
        return str(item)
    return f"{item}.{subitem:02d}"


def format_national_display_code(codigo: str) -> str:
    """
    Formata código nacional como item.subitem.desdobro (2 dígitos cada).
    Ex.: '10301' → '01.03.01'; '100101' → '10.01.01'.
    """
    digits = "".join(ch for ch in (codigo or "") if ch.isdigit())
    if len(digits) < 5:
        return (codigo or "").strip()
    desdobro = int(digits[-2:])
    subitem = int(digits[-4:-2])
    item = int(digits[:-4] or "0")
    return f"{item:02d}.{subitem:02d}.{desdobro:02d}"


def service_catalog_display_label(
    *,
    service_code: str = "",
    codigo_tributacao_nacional_iss: str = "",
    description: str = "",
) -> str:
    raw = (codigo_tributacao_nacional_iss or service_code or "").strip()
    code = format_national_display_code(raw) if raw else ""
    desc = (description or "").strip()
    if code and desc:
        return f"{code} - {desc}"
    return desc or code or service_code or ""


def parse_lista_serv_nac_rows(path: Path, *, sheet_name: str = "LISTA.SERV.NAC.") -> list[dict]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise NationalServiceImportError(
            "Dependência openpyxl ausente. Instale com: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise NationalServiceImportError(
            f"Aba '{sheet_name}' não encontrada. Abas: {', '.join(wb.sheetnames)}"
        )
    ws = wb[sheet_name]
    rows_out: list[dict] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue
        if not row:
            continue
        codigo = _norm_code(row[0] if len(row) > 0 else None)
        if not codigo:
            continue
        item = _int_field(row[1] if len(row) > 1 else None)
        subitem = _int_field(row[2] if len(row) > 2 else None)
        desdobro = _int_field(row[3] if len(row) > 3 else None)
        description = str((row[4] if len(row) > 4 else "") or "").strip()
        if not description:
            continue
        rows_out.append(
            {
                "codigo": codigo[:16],
                "item": item,
                "subitem": subitem,
                "desdobro": desdobro,
                "description": description,
                "lc116_hint": _lc116_hint(item, subitem)[:16],
            }
        )
    wb.close()
    if not rows_out:
        raise NationalServiceImportError(
            "Nenhuma linha com código de tributação nacional encontrada."
        )
    return rows_out


@transaction.atomic
def import_national_service_xlsx(
    *,
    path: Path,
    version_label: str,
    publish: bool = True,
    sheet_name: str = "LISTA.SERV.NAC.",
    notes: str = "",
) -> NationalServiceCatalogVersion:
    label = (version_label or "").strip()
    if not label:
        raise NationalServiceImportError("Informe o rótulo da versão.")
    if NationalServiceCatalogVersion.objects.filter(version_label=label).exists():
        raise NationalServiceImportError(f"Já existe versão com rótulo '{label}'.")

    file_path = Path(path)
    if not file_path.is_file():
        raise NationalServiceImportError(f"Arquivo não encontrado: {file_path}")

    parsed = parse_lista_serv_nac_rows(file_path, sheet_name=sheet_name)
    version = NationalServiceCatalogVersion.objects.create(
        version_label=label,
        source_filename=file_path.name,
        sheet_name=sheet_name,
        status=NationalServiceCatalogVersion.Status.DRAFT,
        notes=notes or "",
        row_count=0,
    )
    NationalServiceItem.objects.bulk_create(
        [
            NationalServiceItem(
                version=version,
                codigo=row["codigo"],
                item=row["item"],
                subitem=row["subitem"],
                desdobro=row["desdobro"],
                description=row["description"],
                lc116_hint=row["lc116_hint"],
            )
            for row in parsed
        ],
        batch_size=500,
    )
    version.row_count = len(parsed)
    version.save(update_fields=["row_count"])

    if publish:
        publish_national_service_version(version)
    return version


@transaction.atomic
def publish_national_service_version(version: NationalServiceCatalogVersion) -> NationalServiceCatalogVersion:
    NationalServiceCatalogVersion.objects.filter(
        status=NationalServiceCatalogVersion.Status.PUBLISHED
    ).exclude(pk=version.pk).update(
        status=NationalServiceCatalogVersion.Status.SUPERSEDED
    )
    version.status = NationalServiceCatalogVersion.Status.PUBLISHED
    version.published_at = timezone.now()
    version.save(update_fields=["status", "published_at"])
    return version


def get_published_national_services():
    version = (
        NationalServiceCatalogVersion.objects.filter(
            status=NationalServiceCatalogVersion.Status.PUBLISHED
        )
        .order_by("-published_at", "-imported_at")
        .first()
    )
    if version is None:
        return version, NationalServiceItem.objects.none()
    return version, version.items.all()


@transaction.atomic
def materialize_national_services_for_tenant(
    *,
    tenant,
    only_missing: bool = True,
) -> dict:
    """
    Copia a lista nacional publicada para ServiceCatalogItem do tenant.
    Assim o campo Serviço da emissão NFS-e (FK) passa a listar os códigos importados.
    service_code = código de tributação nacional (único e estável).
    """
    from apps.master_data.models import ServiceCatalogItem

    version, items = get_published_national_services()
    if version is None:
        raise NationalServiceImportError(
            "Nenhuma Lista de Serviço Nacional publicada. Importe o Anexo B antes."
        )

    created = 0
    updated = 0
    skipped = 0
    for nat in items.iterator():
        existing = ServiceCatalogItem.objects.filter(
            tenant=tenant, service_code=nat.codigo
        ).first()
        if existing is None:
            existing = ServiceCatalogItem.objects.filter(
                tenant=tenant, codigo_tributacao_nacional_iss=nat.codigo
            ).first()

        if existing is None:
            ServiceCatalogItem.objects.create(
                tenant=tenant,
                service_code=nat.codigo,
                description=nat.description,
                lc116_item=nat.lc116_hint,
                codigo_tributacao_nacional_iss=nat.codigo,
                is_active=True,
            )
            created += 1
            continue

        if only_missing:
            skipped += 1
            continue

        existing.description = nat.description
        existing.lc116_item = nat.lc116_hint or existing.lc116_item
        existing.codigo_tributacao_nacional_iss = nat.codigo
        existing.is_active = True
        existing.save(
            update_fields=[
                "description",
                "lc116_item",
                "codigo_tributacao_nacional_iss",
                "is_active",
                "updated_at",
            ]
        )
        updated += 1

    return {
        "version": version.version_label,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total_national": version.row_count,
    }
