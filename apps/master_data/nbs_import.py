"""Importação versionada da Lista NBS — Anexo B NFS-e Nacional (aba LISTA.NBS)."""

from __future__ import annotations

from pathlib import Path

from django.db import transaction
from django.db.models import Q
from django.utils import timezone

from apps.master_data.models import NbsCatalogVersion, NbsItem

NBS_SHEET_CANDIDATES = ("LISTA.NBS_v2.0", "LISTA.NBS", "LISTA NBS")


class NbsImportError(ValueError):
    pass


def normalize_nbs_code(raw: str) -> str:
    digits = "".join(ch for ch in (raw or "") if ch.isdigit())
    return digits[:9]


def format_nbs_display_code(codigo: str) -> str:
    """Formata 9 dígitos como X.XXXX.XX.XX quando possível."""
    digits = normalize_nbs_code(codigo)
    if len(digits) != 9:
        return (codigo or "").strip()
    return f"{digits[0]}.{digits[1:5]}.{digits[5:7]}.{digits[7:9]}"


def _norm_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def _detect_nbs_sheet(wb) -> str:
    for name in NBS_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return name
    for name in wb.sheetnames:
        upper = name.upper()
        if "NBS" in upper:
            return name
    raise NbsImportError(
        f"Nenhuma aba NBS encontrada. Abas: {', '.join(wb.sheetnames)}"
    )


def parse_nbs_rows(path: Path, *, sheet_name: str | None = None) -> tuple[str, list[dict]]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise NbsImportError(
            "Dependência openpyxl ausente. Instale com: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = sheet_name or _detect_nbs_sheet(wb)
    if sheet not in wb.sheetnames:
        wb.close()
        raise NbsImportError(f"Aba '{sheet}' não encontrada.")

    ws = wb[sheet]
    rows_out: list[dict] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue
        if not row:
            continue
        codigo_raw = _norm_cell(row[0] if len(row) > 0 else None)
        codigo = normalize_nbs_code(codigo_raw)
        if len(codigo) != 9:
            continue
        description = str((row[1] if len(row) > 1 else "") or "").strip()
        if not description:
            continue
        rows_out.append({"codigo": codigo, "description": description})
    wb.close()
    if not rows_out:
        raise NbsImportError("Nenhuma linha NBS válida encontrada (código 9 dígitos + descrição).")
    return sheet, rows_out


@transaction.atomic
def import_nbs_xlsx(
    *,
    path: Path,
    version_label: str,
    publish: bool = True,
    sheet_name: str | None = None,
    notes: str = "",
) -> NbsCatalogVersion:
    label = (version_label or "").strip()
    if not label:
        raise NbsImportError("Informe o rótulo da versão.")
    if NbsCatalogVersion.objects.filter(version_label=label).exists():
        raise NbsImportError(f"Já existe versão NBS com rótulo '{label}'.")

    file_path = Path(path)
    if not file_path.is_file():
        raise NbsImportError(f"Arquivo não encontrado: {file_path}")

    resolved_sheet, parsed = parse_nbs_rows(file_path, sheet_name=sheet_name)
    version = NbsCatalogVersion.objects.create(
        version_label=label,
        source_filename=file_path.name,
        sheet_name=resolved_sheet,
        status=NbsCatalogVersion.Status.DRAFT,
        notes=notes or "",
        row_count=0,
    )
    NbsItem.objects.bulk_create(
        [
            NbsItem(
                version=version,
                codigo=row["codigo"],
                description=row["description"],
                is_active=True,
            )
            for row in parsed
        ],
        batch_size=500,
    )
    version.row_count = len(parsed)
    version.save(update_fields=["row_count"])

    if publish:
        publish_nbs_version(version)
    return version


@transaction.atomic
def publish_nbs_version(version: NbsCatalogVersion) -> NbsCatalogVersion:
    NbsCatalogVersion.objects.filter(status=NbsCatalogVersion.Status.PUBLISHED).exclude(
        pk=version.pk
    ).update(status=NbsCatalogVersion.Status.SUPERSEDED)
    version.status = NbsCatalogVersion.Status.PUBLISHED
    version.published_at = timezone.now()
    version.save(update_fields=["status", "published_at"])
    return version


def get_published_nbs_items():
    version = (
        NbsCatalogVersion.objects.filter(status=NbsCatalogVersion.Status.PUBLISHED)
        .order_by("-published_at", "-imported_at")
        .first()
    )
    if version is None:
        return version, NbsItem.objects.none()
    return version, version.items.filter(is_active=True)


def search_nbs(*, query: str = "", limit: int = 20) -> list[dict]:
    """Busca códigos NBS na versão publicada (código prefixo ou descrição)."""
    _version, qs = get_published_nbs_items()
    if _version is None:
        return []

    q = (query or "").strip()
    limit = max(1, min(int(limit or 20), 50))
    if not q:
        return [
            {"codigo": i.codigo, "description": i.description, "display": format_nbs_display_code(i.codigo)}
            for i in qs.order_by("codigo")[:limit]
        ]

    digits = normalize_nbs_code(q)
    if digits:
        qs = qs.filter(Q(codigo__startswith=digits) | Q(codigo=digits))
    else:
        qs = qs.filter(description__icontains=q)

    return [
        {
            "codigo": i.codigo,
            "description": i.description,
            "display": format_nbs_display_code(i.codigo),
        }
        for i in qs.order_by("codigo")[:limit]
    ]


def resolve_nbs_item(*, codigo: str) -> NbsItem | None:
    code = normalize_nbs_code(codigo)
    if len(code) != 9:
        return None
    _version, qs = get_published_nbs_items()
    if _version is None:
        return None
    return qs.filter(codigo=code).first()
